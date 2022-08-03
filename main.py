#!/usr/bin/env python3
# Author: carlos moreno (carlosmorenoaguilera)
# Contributor(s): carlos moreno (carlosmorenoaguilera)
# Jul, 2022

"""
python data scraper for [^1]SANAA calendar data from pdf to excel tabular data



[^1](Honduran Government agency in charge of making water drinkable) 

Dependency: camelot-py==0.10.1 .
"""





import builtins
from PyPDF2 import PdfFileReader
import camelot
from fastapi import File
import pandas as pd
import numpy as np
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
from pandas.io import excel
import openpyxl
import argparse




def generate(file = "HORARIOS AGUA POTABLE-JULIO2022-C3D-RGCC.pdf"):
    #file = ("HORARIOS AGUA POTABLE-JULIO2022-C3D-RGCC.pdf")
    filename ="horario_agua.xlsx"
    pdf = PdfFileReader(open(file,'rb'))
    pages_count =""
    for i in range(pdf.getNumPages()):
        if i != pdf.getNumPages() -1:
            pages_count = pages_count + str(i + 1 ) + ','
        else:
            pages_count = pages_count + str(i+1)
    tables = camelot.read_pdf(file, pages=pages_count)
    print("Total tables extracted (Camelot) ", tables.n)
    tables.export(filename, f="excel", compress=False)
    return tables, filename

def convert(tables):
    filename =  tables[1]
    pages_count = tables[0]#.n
    writer = pd.ExcelWriter('output1.xlsx')
    
    for i in range(pages_count):
        sheet_name = f'page-{i+1}-table-1'
        df = pd.read_excel(filename, sheet_name=sheet_name)
        barrio_header = "Barrio o Colonia"
        colonia_cell_location = ""
        
        for f in range(35):
            current_location = df.at[0,f]
            if ( current_location == barrio_header):
                colonia_cell_location = [0,f]
                break
        for h in range(colonia_cell_location[1]):
            df.drop([h], axis=1, inplace= True)
        
        df.drop([0,1], inplace=True)
            
        if (i == 0): 
            index_count = 0
        else:
            index_count = index_count +  len(df.index)
            
        print("index count output ", index_count)        


        df.to_excel(writer,'Datos', startrow=index_count, index = False, header=False)    



    writer.save()
    writer.close()
    writer.handles = None
    delete_empty_rows('output1.xlsx')
    populate_merged_rows('output1.xlsx')
    print("success generate at: ")



def preparardata():

    cvrt = pd.read_excel('output1.xlsx', sheet_name='Datos')
    wt = pd.ExcelWriter('output2.xlsx')

    headers_list = list(cvrt.columns.values)
    del headers_list[0:3]
    print((headers_list))

    df = cvrt
    df = pd.DataFrame(np.where(df == 'X', 'Y', 'N' ), index=cvrt.index, columns=cvrt.columns)
    df = pd.DataFrame(np.where(df == 'X', cvrt.columns , df ), index=cvrt.index, columns=cvrt.columns)
    df_original = df.assign(dates=pd.Series())

    #join and concat values Y and N 
    df_original["dates"] = df_original[df_original.columns[3:]].apply(lambda x: ', '.join(x.dropna().astype(str)), axis = 1)       
    df_original.drop(df_original.columns[3:-1], axis=1, inplace=True)
    df_original.drop(df_original.columns[0], axis=1, inplace=True)
    df_original = df_original.assign(month=pd.Series())
    df_original['month'] = "Noviembre"
    df_original.reset_index()
    print(df_original)
    df_original.to_excel(wt,sheet_name='Datos',index = False, header=True )
    wt.save()

def prepare_data_for_upload():
    
    #excel_fixed = pd.read_excel('output_fixed.xlsx')
    headers_bulk =  ['Barrio','Hora','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
    excel_bulk =  pd.read_excel('output1.xlsx', sheet_name='Datos', header=None, names=headers_bulk)
    # excel_bulk.drop(columns=excel_bulk.columns[0], axis=1, inplace=True)
    #excel_bulk.drop(index=excel_bulk.index[0:2], axis=0, inplace=True)
    
    
    
    # excel_bulk = pd.DataFrame(np.where(excel_bulk == ' ', 'N', excel_bulk ), index=excel_bulk.index, columns=excel_bulk.columns)
    
    excel_bulk = pd.DataFrame(np.where(excel_bulk == 'X', 'Y', excel_bulk ), index=excel_bulk.index, columns=excel_bulk.columns)
    
    excel_bulk.fillna('N', inplace=True)
    #header_bulk = list(excel_bulk.columns.values)
    #del header_bulk[0:3]        
    #data_frame = pd.DataFrame(np.where(data_frame == 'X', 'Y', 'N'), index=excel_bulk.index, columns=excel_bulk.columns)
    #print(excel_bulk.iloc[:,0])    
    excel_output =  pd.ExcelWriter('Salida.xlsx')
    excel_bulk.to_excel(excel_output,sheet_name='Data', index=False, header=True) #, header=False
    excel_output.save()
    #data_frame = pd.DataFrame(np.where(data_frame == 'X', excel_bulk.columns, data_frame), index=excel_bulk.index, columns=excel_bulk.columns)
    
    #print(data_frame)
    

def uploaddata():
    cred = credentials.Certificate('skarnguiasbase-firebase.json')
    #sanaa_app = firebase_admin.initialize_app()
    firebase_admin.initialize_app(cred, {
        "projectId": "skarnguiasbase"
    })
    db = firestore.client()
    SanaaDF = pd.read_excel('Salida.xlsx', sheet_name='Data')
    SanaaTestDF = SanaaDF   
    dataSet = SanaaTestDF.to_dict("records")    
    data_set = []
    
                
    for item in dataSet:
        a_dict = {}
        a_dict["Dias"] = ''
        for key, value in item.items():
            if key == 'Barrio' or key == 'Hora':
                a_dict[key] = value
            if value == 'Y':
                if a_dict["Dias"] == '':
                    a_dict["Dias"] = key + ","                    
                else:
                    a_dict["Dias"] =  a_dict["Dias"] + key + "," 
                    
        data_set.append(a_dict)    
    

    doc_Collection =  db.collection(u"Sanaa")
    for index, doc in enumerate(data_set):
        print(index,doc)
        doc_ref = db.collection(u"SanaaData").document(str(index))
        doc_ref.set(doc)
        
def get_barrio(barrio):
    cred = credentials.Certificate('skarnguiasbase-firebase.json')
    firebase_admin.initialize_app(cred, {
        "projectId": "skarnguiasbase"
    })
    db = firestore.client()    
    doc_Collection =  db.collection(u"SanaaData")
    result =  doc_Collection.where(u'Barrio', u'==', barrio).stream()
    print(type(result))
    
    for doc in result:
        print(f'{doc.id} => {doc.to_dict()}')

def delete_empty_rows(workbookname):
    rows_delete = [None, '', '']
    wb = openpyxl.load_workbook(workbookname)
    for i in wb.sheetnames:
        print(f"current sheet {i}")
        ws = wb[i]
        column = range(1, ws.max_row)
        for i in reversed(column):
            if ws.cell(i,3).value in rows_delete:
                ws.delete_rows(ws.cell(i,3).row)
        ws.delete_cols(1, 1)
    
    wb.save(workbookname)

def populate_merged_rows(workbookname):
    wb = openpyxl.load_workbook(workbookname)
    for i in wb.sheetnames:
        ws = wb[i]
        column = range(1, ws.max_row)
        for h in column:
            current_row = ws.cell(h,1).value            
            if current_row == None:
                ws.cell(h,1).value = ws.cell(h-1,1).value
    wb.save(workbookname)


def main():
    parser = argparse.ArgumentParser(description=__doc__,         formatter_class=argparse.RawDescriptionHelpFormatter)
    # en construcion 



if __name__ == "__main__":
    main()
    # #pdf_result =  generate("hoja")
    # pdf_result =  [ 5, "horario_agua.xlsx" ]
    # convert(pdf_result)
    # preparardata()
    # prepare_data_for_upload()
    # #uploaddata()
    # #print(get_barrio("CUEVA, MIRADOR, PICACHITO"))
